# coding:utf-8
import os
import logging
import time
import smtplib
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


#当前文件路径
cur_path = os.path.dirname(os.path.realpath(__file__))

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 类名称：InsertLog
# 类的目的：写日志
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：修才华
# 创建时间：2021/11/10
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

# log_path是存放日志的路径
log_path = os.path.join(os.path.dirname(cur_path), 'logs')
# 如果不存在这个logs文件夹，就自动创建一个
if not os.path.exists(log_path):os.mkdir(log_path)

class InsertLog():
    def __init__(self):
        # 文件的命名
        self.logname = os.path.join(log_path, '%s.log'%time.strftime('%Y_%m_%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        # 日志输出格式
        self.formatter = logging.Formatter('[%(asctime)s - %(funcName)s line: %(lineno)3d] - %(levelname)s: %(message)s')

    def __console(self, level, message):
        # 创建一个FileHandler，用于写到本地
        fh = logging.FileHandler(self.logname, 'a')  # 追加模式  这个是python2的
        # fh = logging.FileHandler(self.logname, 'a', encoding='utf-8')  # 这个是python3的
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        # 创建一个StreamHandler,用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.logger.addHandler(ch)

        if level == 'info':
            self.logger.info(message)
        elif level == 'debug':
            self.logger.debug(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        # 这两行代码是为了避免日志输出重复问题
        self.logger.removeHandler(ch)
        self.logger.removeHandler(fh)
        # 关闭打开的文件
        fh.close()

    def debug(self, message):
        self.__console('debug', message)

    def info(self, message):
        self.__console('info', message)

    def warning(self, message):
        self.__console('warning', message)

    def error(self, message):
        self.__console('error', message)

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：GetSkipScripts
# 函数/过程的目的：获取不需要执行的模块名字
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：修才华
# 创建时间：2021/11/10
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

def GetSkipScripts(FilePath):
    try:
        m = []
        wb = load_workbook(FilePath)
        ws = wb['ScriptPath']
        rowcount = ws.max_row
        for i in range(1,rowcount+1):
            cellvalue = ws.cell(row=i,column=3).value
            if cellvalue=='False':
                modulename = ws.cell(row=i,column=2).value
                m.append(modulename)
        wb.close()
        return m
    except BaseException as msg:
        log = InsertLog()
        log.error(msg)

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：GetSkipScripts
# 函数/过程的目的：获取不需要执行的模块名字
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：修才华
# 创建时间：2021/11/10
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

def GetSkipTestCases(FilePath):
    try:
        t = []
        wb = load_workbook(FilePath)
        sheels = wb.sheetnames
        #print sheels
        for i in sheels:
            ws = wb[i]
            rowcount = ws.max_row
            for j in range(1,rowcount+1):
                cellvalue = ws.cell(row=j,column=7).value
                if cellvalue=='False':
                    testcasename = ws.cell(row=j,column=1).value
                    t.append(testcasename)
        wb.close()
        return t
    except BaseException as msg:
        log = InsertLog()
        log.error(msg)



#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：GetNewReport
# 函数/过程的目的：获取最新报告文件
# 假设：无
# 影响：无
# 输入：无
# 返回值：文件全路径
# 创建者：修才华
# 创建时间：2021/11/10
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
FD = "C://Users//Andrew//PycharmProjects//FramWork//report//"
def GetNewReport(FileDir=FD):
    #打印目录所在所有文件名（列表对象）
    l = os.listdir(FileDir)
    l.sort(key=lambda fn:os.path.getmtime(FileDir + "\\" + fn))
    f = os.path.join(FileDir,l[-1])
    return f

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 类名称：SendEmail
# 类的目的：发送文本邮件或发送带附件邮件
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：修才华
# 创建时间：2021/11/10
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
def SendEmail(sender, psw, receiver, smtpserver, report_file, port):
    with open(report_file, "rb") as f:
        mail_body = f.read()
    # 定义邮件内容
    msg = MIMEMultipart()
    body = MIMEText(mail_body, _subtype='html', _charset='utf-8')
    msg['Subject'] = u"自动化测试报告"
    msg["from"] = sender
    msg["to"] = psw
    msg.attach(body)
    # 添加附件
    att = MIMEText(open(report_file, "rb").read(), "base64", "utf-8")
    att["Content-Type"] = "application/octet-stream"
    att["Content-Disposition"] = 'attachment; filename= "report.html"'
    msg.attach(att)
    try:
        smtp = smtplib.SMTP_SSL(smtpserver, port)
    except:
        smtp = smtplib.SMTP()
        smtp.connect(smtpserver, port)
    # 用户名,客户端授权密码%(module)s
    smtp.login(sender, psw)
    smtp.sendmail(sender, receiver, msg.as_string())
    smtp.quit()
    print (u'邮件发送成功！')