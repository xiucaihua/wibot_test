#coding:utf-8
# 导入需要用到的相关包
import openpyxl
from openpyxl import load_workbook


#第二个方法
class Excel():

    def __init__(self, file,index):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[index]
        self.ws = self.wb[self.sheet]

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(2, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    def greepData(self,listGroup):
        pass

class GreepData():

    # 合并单元格的操作
    def mergeCell(self, file, savefile):
        workbook = load_workbook(file)
        print(workbook.sheetnames)
        sheet = workbook.active
        sheet.merge_cells('A2:A19')
        sheet.merge_cells(start_row=2, start_column=1, end_row=19, end_column=1)
        cell = sheet['A2']
        cell.value = '腾信国际物流'

        sheet.merge_cells('A20:A20')
        sheet.merge_cells(start_row=20, start_column=1, end_row=20, end_column=1)
        cell = sheet['A20']
        cell.value = '企点采点助手'

        sheet.merge_cells('A21:A23')
        sheet.merge_cells(start_row=21, start_column=1, end_row=23, end_column=1)
        cell = sheet['A21']
        cell.value = '建行杭州之江支行'

        sheet.merge_cells('A24:A28')
        sheet.merge_cells(start_row=24, start_column=1, end_row=28, end_column=1)
        cell = sheet['A24']
        cell.value = '北银天津'

        sheet.merge_cells('A29:A29')
        sheet.merge_cells(start_row=29, start_column=1, end_row=29, end_column=1)
        cell = sheet['A29']
        cell.value = '江北支行'
        workbook.save(savefile)

        # 合并单元格的操作

    def mergeAssistantCell(self, file, savefile):
        workbook = load_workbook(file)
        print(workbook.sheetnames)
        sheet = workbook.active

        sheet.merge_cells('A4:A6')
        sheet.merge_cells(start_row=4, start_column=1, end_row=6, end_column=1)
        cell = sheet['A4']
        cell.value = '建行杭州之江支行'

        sheet.merge_cells('A7:A11')
        sheet.merge_cells(start_row=7, start_column=1, end_row=11, end_column=1)
        cell = sheet['A7']
        cell.value = '北银天津'

        sheet.merge_cells('A12:A14')
        sheet.merge_cells(start_row=12, start_column=1, end_row=14, end_column=1)
        cell = sheet['A12']
        cell.value = '合川支行'

        sheet.merge_cells('A15:A17')
        sheet.merge_cells(start_row=15, start_column=1, end_row=17, end_column=1)
        cell = sheet['A15']
        cell.value = '江北支行'
        workbook.save(savefile)


    """
    根据非真实的群组信息，进行删除，然后保存真实的群组信息操作。
    """
    def deleteRow(self,fileName,SheetName,listData,savaFile):
        wb = openpyxl.load_workbook(fileName)
        sht = wb[SheetName]

        c = 2
        for r in range(sht.max_row, 0, -1):
            for a in listData:
                if sht.cell(r, c).value == a:
                    sht.delete_rows(r)
        wb.save(savaFile)
        wb.close()

    """
    获取到非真实的群组信息
    """
    def getDifferent(self,fileName,index):
        listGroup = ['💝建行之江支行福利优惠群2🎊', '💝建行之江支行福利优惠群✨', '💝建行之江支行福利优惠群3✨',
                     '小冰老师的1群', '河西支行 厅堂咨询2群', '河西支行 微信营销群',
                     '京彩生活', '河西支行 厅堂咨询群🌷🌟', '恒丰银行江北支行福利3群',
                     '云翔国际——腾信国际', '祺运达国际——腾信国际', '速特国际&腾信国际',
                     '橙辉～腾信', 'F安达国际一腾信国际', '腾信国际——F拓普森',
                     '德华供应链—深圳腾信国际', '腾信国际---广州国运', 'F欧顺通跨境-腾信国际',
                     '飞莱F--腾信国际', '山海国际=腾信', '承磊=腾信',
                     '腾信=华宣', '捷讯通—腾信', '欧顺通—深圳腾信国际',
                     '信必达&腾信国际', '时信国际-腾信国际', '腾信美线空运清关', '🌞优质电子元器件供应商群']

        #真实的客户群组信息
        excel = Excel(fileName,index) #/群人数在线统计综合.xlsx
        #获取系统上所有的群组信息
        listcolumndata = excel.getColValues(2)
        # print("所有的：", listcolumndata)
        # print("系统素有群组的值大小为：", len(listcolumndata))
        # print("客户真实的群组大小为：", len(listGroup))

        listRsulet = []
        for a in listcolumndata:
            if a not in listGroup:
                listRsulet.append(a)
        #print(listRsulet)
        return listRsulet

#if __name__ == '__main__':

    #设置合并单元格8
    # greepData=GreepData()
    # greepData.mergeCell("./删除后保留的数据.xlsx","./合并单元格的.xlsx")
    #
    # #用例7
    # greepData=GreepData()
    # listData = greepData.getDifferent('./群人数在线统计综合.xlsx', 0)
    # greepData.deleteRow('./群人数在线统计综合.xlsx','Sheet1',listData,'./删除后保留的数据.xlsx')

    #设置第一第二列宽度9
    # sheetClass=SheetClass()
    # sheetClass.setExcelStyle('./合并单元格的.xlsx','Sheet1','./在群人数最终版本数据.xlsx')